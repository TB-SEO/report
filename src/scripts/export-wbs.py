# -*- coding: utf-8 -*-
"""Export WBS xlsx → wbs-data.json with full column/row outline ranges."""
import json
import re
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path
from zipfile import ZipFile

import openpyxl
from openpyxl.utils import get_column_letter

candidates = [
    Path(r"C:\Users\AM-DBD-LT001\Downloads\SEO 전략 기획 (1).xlsx"),
    Path(r"C:\develop_folder\SEO-Report_server\data\wbs\seo-strategy.xlsx"),
    Path(r"C:\Users\AM-DBD-LT001\Downloads\SEO 전략 기획 (1) - 복사본.xlsx"),
]
src = next(p for p in candidates if p.exists())
root = Path(__file__).resolve().parents[2]
dest_xlsx = root / "data" / "wbs" / "seo-strategy.xlsx"
dest_json = root / "src" / "web" / "public" / "wbs-data.json"
dest_xlsx.parent.mkdir(parents=True, exist_ok=True)
shutil.copy2(src, dest_xlsx)

wb = openpyxl.load_workbook(dest_xlsx, data_only=True)


def rgb(color):
    if not color:
        return None
    if getattr(color, "type", None) == "rgb" and color.rgb and str(color.rgb) not in ("00000000", "None"):
        text = str(color.rgb)
        return "#" + text[-6:] if len(text) >= 6 else text
    return None


def val(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def sheet_target_map(z: ZipFile):
    wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    sheets = re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml)
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rid_to = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))
    out = {}
    for name, rid in sheets:
        target = rid_to[rid].replace("\\", "/")
        if not target.startswith("xl/"):
            target = "xl/" + target.lstrip("/")
        out[name] = target
    return out


def parse_col_outline(xml: str, limit_c: int):
    """Expand OOXML <col min max outlineLevel collapsed/> onto each column index."""
    items = [{"level": 0, "collapsed": False, "hidden": False} for _ in range(limit_c)]
    for m in re.finditer(
        r"<[\w:]*col\b([^>]*)/?>",
        xml,
    ):
        attrs = m.group(1)
        amin = re.search(r'\bmin="(\d+)"', attrs)
        amax = re.search(r'\bmax="(\d+)"', attrs)
        if not amin or not amax:
            continue
        lo = int(amin.group(1))
        hi = int(amax.group(1))
        lvl = int(re.search(r'\boutlineLevel="(\d+)"', attrs).group(1)) if re.search(r'\boutlineLevel="(\d+)"', attrs) else 0
        collapsed = 'collapsed="1"' in attrs or 'collapsed="true"' in attrs
        hidden = 'hidden="1"' in attrs or 'hidden="true"' in attrs
        for c in range(lo, hi + 1):
            if 1 <= c <= limit_c:
                items[c - 1] = {"level": lvl, "collapsed": collapsed, "hidden": hidden}
    return items


def parse_row_outline(xml: str, limit_r: int):
    items = [{"level": 0, "collapsed": False, "hidden": False} for _ in range(limit_r)]
    for m in re.finditer(r"<[\w:]*row\b([^>]*)>", xml):
        attrs = m.group(1)
        rr = re.search(r'\br="(\d+)"', attrs)
        if not rr:
            continue
        r = int(rr.group(1))
        if r < 1 or r > limit_r:
            continue
        lvl = int(re.search(r'\boutlineLevel="(\d+)"', attrs).group(1)) if re.search(r'\boutlineLevel="(\d+)"', attrs) else 0
        collapsed = 'collapsed="1"' in attrs or 'collapsed="true"' in attrs
        hidden = 'hidden="1"' in attrs or 'hidden="true"' in attrs
        items[r - 1] = {"level": lvl, "collapsed": collapsed, "hidden": hidden}
    return items


def outline_pr(xml: str):
    m = re.search(r"<[\w:]*outlinePr\b([^>]*)/?>", xml)
    if not m:
        return False, False
    attrs = m.group(1)
    below = 'summaryBelow="1"' in attrs or 'summaryBelow="true"' in attrs
    right = 'summaryRight="1"' in attrs or 'summaryRight="true"' in attrs
    return below, right


with ZipFile(dest_xlsx) as zf:
    targets = sheet_target_map(zf)
    sheet_xml = {name: zf.read(target).decode("utf-8") for name, target in targets.items()}

out = {"sheets": []}
for name in wb.sheetnames:
    ws = wb[name]
    max_r, max_c = ws.max_row, ws.max_column
    limit_r = min(max_r or 1, 200)
    limit_c = min(max_c or 1, 160)
    xml = sheet_xml.get(name, "")
    summary_below, summary_right = outline_pr(xml)
    if name in ("WBS", "Content") and not summary_below and not summary_right:
        # Excel default often omitted; WBS screenshot uses summary left / above
        summary_below, summary_right = False, False
    cols_ol = parse_col_outline(xml, limit_c) if xml else [
        {"level": 0, "collapsed": False, "hidden": False} for _ in range(limit_c)
    ]
    rows_ol = parse_row_outline(xml, limit_r) if xml else [
        {"level": 0, "collapsed": False, "hidden": False} for _ in range(limit_r)
    ]
    widths = []
    for i in range(1, limit_c + 1):
        width = ws.column_dimensions[get_column_letter(i)].width
        widths.append(round(float(width), 2) if width else 8)
    rows = []
    for r in range(1, limit_r + 1):
        row = []
        for c in range(1, limit_c + 1):
            cell = ws.cell(r, c)
            item = {"v": val(cell.value)}
            fill = None
            if cell.fill and cell.fill.patternType and cell.fill.patternType != "none":
                fill = rgb(cell.fill.fgColor)
            if fill and fill not in ("#000000",):
                item["fill"] = fill
            font = cell.font
            if font:
                color = rgb(font.color)
                if color and color not in ("#000000",):
                    item["color"] = color
                if font.bold:
                    item["b"] = True
                if font.size:
                    item["sz"] = font.size
                if font.underline and font.underline != "none":
                    item["u"] = True
            align = cell.alignment
            if align:
                if align.horizontal:
                    item["h"] = align.horizontal
                if align.wrap_text:
                    item["wrap"] = True
            row.append(item)
        rows.append(row)
    out["sheets"].append(
        {
            "name": name,
            "maxRow": limit_r,
            "maxCol": limit_c,
            "widths": widths,
            "merges": [str(m) for m in ws.merged_cells.ranges],
            "freeze": str(ws.freeze_panes) if ws.freeze_panes else None,
            "outline": {
                "summaryBelow": summary_below,
                "summaryRight": summary_right,
                "rows": rows_ol,
                "cols": cols_ol,
            },
            "rows": rows,
        }
    )

dest_json.write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")
# quick verify
wbs = next(s for s in out["sheets"] if s["name"] == "WBS")
lv = [c["level"] for c in wbs["outline"]["cols"]]
print("ok", dest_json)
print("src", src)
print("col levels nonzero", sum(1 for x in lv if x), "max", max(lv), "sample", "".join(str(x) for x in lv[0:50]))

import subprocess

sync = subprocess.run(["npm", "run", "db:sync"], cwd=root)
if sync.returncode:
    print("WBS JSON은 저장했습니다. npm run db:sync 로 Supabase에 올리세요.")
